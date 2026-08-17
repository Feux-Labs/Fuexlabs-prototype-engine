import openpyxl
import re
import os
import csv
import urllib.parse

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(BASE_DIR, "prototype_generator", "template.html")
OUT_DIR = os.path.join(BASE_DIR, "prototypes")
CSV_OUT_PATH = os.path.join(BASE_DIR, "prototype_links.csv")

DEFAULT_XLSX_PATHS = [
    os.path.join(BASE_DIR, "Feux_Labs_Lead_Pipeline.xlsx"),
    os.path.join(os.path.dirname(BASE_DIR), "Feux_Labs_Lead_Pipeline.xlsx"),
    "/home/claude/Feux_Labs_Lead_Pipeline.xlsx"
]

COLORS = [
    ("#C99700", "#8A6900"),  # gold
    ("#2F6FED", "#1E4FB0"),  # blue
    ("#1B7A3D", "#0F5228"),  # green
    ("#B0442C", "#7A2E1D"),  # rust
    ("#5B4FCF", "#3E36A0"),  # violet
    ("#0E7C7B", "#0A5958"),  # teal
]

def slugify(name):
    s = name.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    return s[:60]

def initials(name):
    words = [w for w in re.split(r"\s+", name) if w and w.lower() not in ("of", "the", "and", "&", "for")]
    letters = [w[0].upper() for w in words[:2]]
    return "".join(letters) if letters else name[:2].upper()

def wa_link(phone, school_name, proto_url):
    if not phone:
        return ""
    digits = re.sub(r"[^\d]", "", phone)
    if digits.startswith("0"):
        digits = "234" + digits[1:]
    msg = (f"Hi, this is Feux Labs. We're a tech studio in Abuja and we build websites & apps. "
           f"We put together a free concept preview of what a website for {school_name} could look "
           f"like: {proto_url} — take a look and let us know if you'd like the real version built.")
    return f"https://wa.me/{digits}?text={urllib.parse.quote(msg)}"

import sys

def find_excel_file():
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        return sys.argv[1]
    for p in DEFAULT_XLSX_PATHS:
        if os.path.exists(p):
            return p
    return None

def main():
    xlsx_path = find_excel_file()
    if not xlsx_path:
        print(f"Error: Excel file 'Feux_Labs_Lead_Pipeline.xlsx' not found!")
        print(f"Usage: python prototype_generator/generate.py <path_to_excel_file>")
        return

    print(f"Loading Excel lead data from: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    leads = wb["Leads"]

    with open(TEMPLATE_PATH, encoding="utf-8") as f:
        template = f.read()

    rows_out = []
    color_i = 0

    for r in range(13, 104):
        name = leads.cell(row=r, column=1).value
        if not name:
            continue
        problem = leads.cell(row=r, column=9).value or ""
        if "NO WEBSITE" not in problem and "no website" not in problem.lower():
            continue

        phone = leads.cell(row=r, column=4).value or ""
        notes = leads.cell(row=r, column=13).value or ""
        address = notes.split("|")[0].strip() if notes else "Abuja, Nigeria"

        slug = slugify(name)
        accent, accent_dark = COLORS[color_i % len(COLORS)]
        color_i += 1

        html = template
        html = html.replace("{{SCHOOL_NAME}}", str(name))
        html = html.replace("{{INITIALS}}", initials(str(name)))
        html = html.replace("{{COLOR_DARK}}", accent_dark)
        html = html.replace("{{COLOR}}", accent)
        html = html.replace("{{TAGLINE}}", f"Nurturing young minds in {address.split(',')[0] if address else 'Abuja'}.")
        html = html.replace(
            "{{ABOUT_TEXT}}",
            f"{name} is committed to providing quality education and a supportive learning "
            f"environment for every student."
        )
        html = html.replace("{{ADDRESS}}", address)
        html = html.replace("{{PHONE}}", str(phone) if phone else "Contact via WhatsApp")

        school_dir = os.path.join(OUT_DIR, slug)
        os.makedirs(school_dir, exist_ok=True)
        with open(os.path.join(school_dir, "index.html"), "w", encoding="utf-8") as f:
            f.write(html)

        proto_url = f"https://prototypes.feuxlabs.com.ng/{slug}/"
        rows_out.append({
            "school": name,
            "slug": slug,
            "phone": phone,
            "prototype_url": proto_url,
            "whatsapp_send_link": wa_link(str(phone), str(name), proto_url),
        })

    with open(CSV_OUT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["school", "slug", "phone", "prototype_url", "whatsapp_send_link"])
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"Successfully generated {len(rows_out)} prototype sites into {OUT_DIR}")
    print(f"Master link CSV saved to {CSV_OUT_PATH}")
    return rows_out

if __name__ == "__main__":
    main()
